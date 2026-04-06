"""
run_full.py  v2.0  –  Tam Pipeline Çalıştırıcı
================================================
Akakçe scraper → PostgreSQL DB → Excel / CSV raporlama

Kullanım:
  python run_full.py --url "https://www.akakce.com/cep-telefonu.html" --max 100
  python run_full.py --url "https://www.akakce.com/laptop.html" --max 200 --no-db
  python run_full.py --categories (bilinen tüm kategorileri çek)
"""

from __future__ import annotations

import argparse
import asyncio
import json
import logging
import os
import sys
import time
from datetime import datetime
from pathlib import Path
from typing import Optional

# ── Çıktı dizinleri ──────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
OUTPUT_DIR  = BASE_DIR / "outputs"
IMAGE_DIR   = OUTPUT_DIR / "images"
COOKIE_DIR  = OUTPUT_DIR / "cookies"
LOG_DIR     = OUTPUT_DIR / "logs"

for d in (OUTPUT_DIR, IMAGE_DIR, COOKIE_DIR, LOG_DIR):
    d.mkdir(parents=True, exist_ok=True)

# ── Logging ayarları ─────────────────────────────────────────────────────────
ts = datetime.now().strftime("%Y%m%d_%H%M%S")
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(name)s: %(message)s",
    handlers=[
        logging.StreamHandler(sys.stdout),
        logging.FileHandler(LOG_DIR / f"run_{ts}.log", encoding="utf-8"),
    ],
)
log = logging.getLogger("run_full")

# ── Bilinen kategoriler ───────────────────────────────────────────────────────
KNOWN_CATEGORIES = {
    "cep-telefonu":  "https://www.akakce.com/cep-telefonu.html",
    "laptop":        "https://www.akakce.com/laptop,bilgisayar.html",
    "tablet":        "https://www.akakce.com/tablet.html",
    "tv":            "https://www.akakce.com/televizyon.html",
    "kulaklik":      "https://www.akakce.com/kulaklik.html",
    "saat":          "https://www.akakce.com/saat.html",
    "kamera":        "https://www.akakce.com/kamera.html",
    "monitor":       "https://www.akakce.com/monitor.html",
}


# ─────────────────────────────────────────────────────────────────────────────
# DB KAYIT KATMANI  (DB yoksa sessizce atla)
# ─────────────────────────────────────────────────────────────────────────────

async def save_results_to_db(result: dict) -> dict[str, int]:
    """
    Scraper sonuçlarını PostgreSQL'e yazar.
    Returns: {"products": N, "sellers": N, "images": N, "errors": N}
    """
    stats = {"products": 0, "sellers": 0, "images": 0, "errors": 0}
    try:
        from database import (
            AsyncSessionLocal,
            bulk_upsert_products,
            bulk_insert_price_history,
            get_or_create_brand,
            get_or_create_category,
        )
        from models import Store, SellerPrice, ProductImage
        from sqlalchemy import select
    except ImportError as exc:
        log.warning("DB modülleri yüklenemedi, DB kaydı atlandı: %s", exc)
        return stats

    products = result.get("products", [])
    if not products:
        return stats

    category_url = result.get("category_url", "")
    category_slug = _url_to_slug(category_url)

    async with AsyncSessionLocal() as session:
        async with session.begin():
            try:
                # Kategori
                cat_id = await get_or_create_category(
                    session,
                    name=category_slug.replace("-", " ").title(),
                    slug=category_slug,
                    url=category_url,
                )

                # Ürünleri toplu upsert
                product_rows = []
                brand_cache: dict[str, int] = {}

                for p in products:
                    brand_name = p.get("brand") or ""
                    brand_id = None
                    if brand_name:
                        if brand_name not in brand_cache:
                            brand_cache[brand_name] = await get_or_create_brand(
                                session, name=brand_name,
                                slug=brand_name.lower().replace(" ", "-"),
                            )
                        brand_id = brand_cache[brand_name]

                    product_rows.append({
                        "external_id":  str(p.get("id", "")),
                        "name":         p.get("name", ""),
                        "url":          p.get("url", ""),
                        "price":        p.get("price"),
                        "old_price":    p.get("old_price"),
                        "image_url":    p.get("image_url"),
                        "brand_id":     brand_id,
                        "category_id":  cat_id,
                        "in_stock":     p.get("in_stock", True),
                        "specs":        p.get("specs") or {},
                    })

                id_map = await bulk_upsert_products(session, product_rows)
                stats["products"] = len(id_map)

                # Fiyat geçmişi
                ph_rows = [
                    {"product_id": db_id, "price": row["price"], "source": "akakce"}
                    for row, db_id in id_map.items()
                    if row.get("price")
                ]
                if ph_rows:
                    await bulk_insert_price_history(session, ph_rows)

                # Satıcı fiyatları & görseller (varsa)
                for p, db_product_id in id_map.items():
                    for sp in (p.get("seller_prices") or []):
                        store_id = await _get_or_create_store(
                            session, sp.get("store_name", ""), sp.get("store_logo_url")
                        )
                        seller = SellerPrice(
                            product_id=db_product_id,
                            store_id=store_id,
                            price=sp.get("price", 0),
                            url=sp.get("url"),
                            is_best=sp.get("is_best", False),
                            in_stock=sp.get("in_stock", True),
                        )
                        session.add(seller)
                        stats["sellers"] += 1

                    for idx, img_url in enumerate(p.get("images") or []):
                        img = ProductImage(
                            product_id=db_product_id,
                            url=img_url,
                            sort_order=idx,
                            is_main=(idx == 0),
                        )
                        session.add(img)
                        stats["images"] += 1

            except Exception as exc:
                log.error("DB kayıt hatası: %s", exc, exc_info=True)
                stats["errors"] += 1
                raise

    return stats


async def _get_or_create_store(session, name: str, logo_url: Optional[str] = None) -> Optional[int]:
    """Store tablosunda mağazayı bul veya oluştur."""
    if not name:
        return None
    try:
        from models import Store
        from sqlalchemy import select
        result = await session.execute(select(Store).where(Store.name == name))
        store = result.scalar_one_or_none()
        if store:
            return store.id
        store = Store(name=name, logo_url=logo_url, is_active=True)
        session.add(store)
        await session.flush()
        return store.id
    except Exception:
        return None


def _url_to_slug(url: str) -> str:
    """https://www.akakce.com/cep-telefonu.html → cep-telefonu"""
    try:
        part = url.rstrip("/").split("/")[-1]
        return part.replace(".html", "").replace(",", "-")
    except Exception:
        return "unknown"


# ─────────────────────────────────────────────────────────────────────────────
# EXCEL / CSV RAPOR
# ─────────────────────────────────────────────────────────────────────────────

def save_report(result: dict, output_dir: Path = OUTPUT_DIR) -> Path:
    """Scraper sonuçlarını Excel (3 sheet) + CSV olarak kaydeder."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    products      = result.get("products", [])
    seller_prices = []
    filters_data  = result.get("filters", [])

    for p in products:
        for sp in (p.get("seller_prices") or []):
            seller_prices.append({
                "product_id":   p.get("id"),
                "product_name": p.get("name"),
                "store":        sp.get("store_name", ""),
                "price":        sp.get("price"),
                "is_best":      sp.get("is_best", False),
                "in_stock":     sp.get("in_stock", True),
                "url":          sp.get("url", ""),
            })

    slug     = _url_to_slug(result.get("category_url", "unknown"))
    ts_str   = datetime.now().strftime("%Y%m%d_%H%M%S")
    xlsx_path = output_dir / f"akakce_{slug}_{ts_str}.xlsx"
    csv_path  = output_dir / f"akakce_{slug}_{ts_str}.csv"

    wb = openpyxl.Workbook()

    # Sheet 1: Ürünler
    ws1 = wb.active
    ws1.title = "Urunler"
    prod_headers = ["ID", "Isim", "Marka", "Fiyat (TL)", "Eski Fiyat", "Indirim %",
                    "Stok", "URL", "Gorsel", "Scrape Zamani"]
    _write_header_row(ws1, prod_headers)
    for p in products:
        price     = p.get("price")
        old_price = p.get("old_price")
        drop_pct  = None
        if price and old_price and old_price > price:
            drop_pct = round((old_price - price) / old_price * 100, 1)
        ws1.append([p.get("id",""), p.get("name",""), p.get("brand",""),
                    price, old_price, drop_pct,
                    "Evet" if p.get("in_stock", True) else "Hayir",
                    p.get("url",""), p.get("image_url",""), p.get("scraped_at","")])
    _auto_width(ws1)

    # Sheet 2: Satici Fiyatlari
    ws2 = wb.create_sheet("Satici Fiyatlari")
    _write_header_row(ws2, ["Urun ID", "Urun Adi", "Magaza", "Fiyat (TL)", "En Iyi?", "Stok", "URL"])
    for sp in seller_prices:
        ws2.append([sp["product_id"], sp["product_name"], sp["store"], sp["price"],
                    "Evet" if sp["is_best"] else "", "Evet" if sp["in_stock"] else "Hayir", sp["url"]])
    _auto_width(ws2)

    # Sheet 3: Filtreler
    ws3 = wb.create_sheet("Filtreler")
    _write_header_row(ws3, ["Grup", "Filtre Degeri", "Akakce ID", "URL"])
    for f in filters_data:
        ws3.append([f.get("group",""), f.get("value",""), f.get("akakce_id",""), f.get("url","")])
    _auto_width(ws3)

    wb.save(xlsx_path)
    log.info("Excel kaydedildi: %s", xlsx_path)

    # CSV
    import csv
    with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "id","name","brand","price","old_price","drop_pct","in_stock","url","image_url","scraped_at"])
        writer.writeheader()
        for p in products:
            price     = p.get("price")
            old_price = p.get("old_price")
            drop_pct  = round((old_price - price) / old_price * 100, 1) if (price and old_price and old_price > price) else None
            writer.writerow({"id": p.get("id",""), "name": p.get("name",""), "brand": p.get("brand",""),
                             "price": price, "old_price": old_price, "drop_pct": drop_pct,
                             "in_stock": p.get("in_stock",True), "url": p.get("url",""),
                             "image_url": p.get("image_url",""), "scraped_at": p.get("scraped_at","")})
    log.info("CSV kaydedildi: %s", csv_path)
    return xlsx_path


def _write_header_row(ws, headers):
    from openpyxl.styles import Font, PatternFill, Alignment
    ws.append(headers)
    fill = PatternFill("solid", fgColor="1F4E79")
    font = Font(bold=True, color="FFFFFF", name="Calibri")
    for cell in ws[1]:
        cell.fill = fill; cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28


def _auto_width(ws, min_w=10, max_w=50):
    from openpyxl.utils import get_column_letter
    for i, col in enumerate(ws.columns, 1):
        ml = max((len(str(c.value)) if c.value else 0 for c in col), default=0)
        ws.column_dimensions[get_column_letter(i)].width = max(min_w, min(max_w, ml + 4))


# ─────────────────────────────────────────────────────────────────────────────
# GÖRSEL İNDİRME
# ─────────────────────────────────────────────────────────────────────────────

def download_images(products: list, image_dir: Path = IMAGE_DIR) -> int:
    import urllib.request, hashlib
    downloaded = 0
    for p in products:
        urls = []
        if p.get("image_url"):
            urls.append(p["image_url"])
        urls.extend(p.get("images") or [])
        for url in urls:
            if not url or not url.startswith("http"):
                continue
            ext  = url.split(".")[-1].split("?")[0][:4] or "jpg"
            name = hashlib.md5(url.encode()).hexdigest()[:16] + "." + ext
            dest = image_dir / name
            if dest.exists():
                continue
            try:
                req = urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"})
                with urllib.request.urlopen(req, timeout=10) as resp:
                    dest.write_bytes(resp.read())
                downloaded += 1
            except Exception as exc:
                log.debug("Gorsel indirilemedi %s: %s", url, exc)
    log.info("%d gorsel indirildi", downloaded)
    return downloaded


# ─────────────────────────────────────────────────────────────────────────────
# ANA PIPELINE
# ─────────────────────────────────────────────────────────────────────────────

def run_pipeline(url, max_products=100, max_pages=10,
                 use_db=True, download_imgs=True, headless=True, detail=True) -> dict:
    log.info("=" * 70)
    log.info("Pipeline: %s  max=%d  db=%s  img=%s", url, max_products, use_db, download_imgs)
    log.info("=" * 70)
    t0 = time.monotonic()

    # 1. Scrape
    try:
        from scraper_v4 import AkakceScraperV4
        with AkakceScraperV4(headless=headless, download_images=False, detail_workers=3) as scraper:
            result = scraper.scrape_category(url, max_products=max_products,
                                             max_pages=max_pages, detail=detail)
    except Exception as exc:
        log.error("Scraper hatasi: %s", exc, exc_info=True)
        return {"error": str(exc), "products": [], "category_url": url}

    products = result.get("products", [])
    log.info("Scraping tamam: %d urun  %.1fs", len(products), time.monotonic() - t0)

    # 2. Excel / CSV
    try:
        result["xlsx_path"] = str(save_report(result))
    except Exception as exc:
        log.error("Rapor hatasi: %s", exc, exc_info=True)

    # 3. Gorseller
    if download_imgs and products:
        try:
            result["images_downloaded"] = download_images(products, IMAGE_DIR)
        except Exception as exc:
            log.error("Gorsel indirme hatasi: %s", exc, exc_info=True)

    # 4. DB
    if use_db and products:
        try:
            db_stats = asyncio.run(save_results_to_db(result))
            result["db_stats"] = db_stats
            log.info("DB kayit: %s", db_stats)
        except Exception as exc:
            log.warning("DB kayit basarisiz: %s", exc)
            result["db_stats"] = {"error": str(exc)}

    total_sec = time.monotonic() - t0
    summary = {
        "url": url,
        "products_scraped": len(products),
        "filters_found":    len(result.get("filters", [])),
        "total_time_sec":   round(total_sec, 1),
        "xlsx":             result.get("xlsx_path", ""),
        "db":               result.get("db_stats", {}),
        "images":           result.get("images_downloaded", 0),
    }
    log.info("OZET: %s", json.dumps(summary, ensure_ascii=False, indent=2))
    return {**result, **summary}


# ─────────────────────────────────────────────────────────────────────────────
# CLI
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="Akakce Scraper – Tam Pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Ornekler:
  python run_full.py --url https://www.akakce.com/cep-telefonu.html --max 100
  python run_full.py --category cep-telefonu --max 200 --no-db
  python run_full.py --categories   # tum bilinen kategoriler
""",
    )
    parser.add_argument("--url",         type=str)
    parser.add_argument("--category",    type=str, choices=list(KNOWN_CATEGORIES))
    parser.add_argument("--categories",  action="store_true")
    parser.add_argument("--max",         type=int, default=100)
    parser.add_argument("--max-pages",   type=int, default=10)
    parser.add_argument("--no-db",       action="store_true")
    parser.add_argument("--no-images",   action="store_true")
    parser.add_argument("--no-headless", action="store_true")
    parser.add_argument("--no-detail",   action="store_true")
    args = parser.parse_args()

    if args.categories:
        targets = list(KNOWN_CATEGORIES.values())
    elif args.category:
        targets = [KNOWN_CATEGORIES[args.category]]
    elif args.url:
        targets = [args.url]
    else:
        parser.print_help(); sys.exit(1)

    all_results = []
    for target_url in targets:
        r = run_pipeline(
            url=target_url, max_products=args.max, max_pages=args.max_pages,
            use_db=not args.no_db, download_imgs=not args.no_images,
            headless=not args.no_headless, detail=not args.no_detail,
        )
        all_results.append(r)
        if len(targets) > 1:
            log.info("Kategori arasi bekleniyor (15s)…")
            time.sleep(15)

    if len(all_results) > 1:
        log.info("GENEL OZET: %d kategori, toplam %d urun",
                 len(all_results), sum(r.get("products_scraped", 0) for r in all_results))


if __name__ == "__main__":
    if sys.platform == "win32":
        asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
    main()
